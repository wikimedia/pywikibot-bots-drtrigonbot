# -*- coding: utf-8  -*-
"""
Robot which will does substitutions of tags within wiki page content with external or
other wiki text data. Like dynamic text updating.

Look at http://de.wikipedia.org/wiki/Benutzer:DrTrigon/Benutzer:DrTrigonBot/config.css
for 'postproc' example code.

Look at https://wiki.toolserver.org/view/Mail how to setup mail handling. The
following code was used in file '$HOME/.forward+subster':
--- --- --- --- --- --- --- --- --- --- 
> ~/data/subster/mail_inbox
--- --- --- --- --- --- --- --- --- --- 
in order to enable mail (mbox style) storage in given file for address:
drtrigon+subster@toolserver.org
"""
## @package subster
#  @brief   Dynamic Text Substitutions Robot
#
#  @copyright Dr. Trigon, 2009-2010
#
#  @section FRAMEWORK
#
#  Python wikipedia robot framework, DrTrigonBot.
#  @see http://pywikipediabot.sourceforge.net/
#  @see http://de.wikipedia.org/wiki/Benutzer:DrTrigonBot
#
#  External code / other modules used are listed here.
#  @see https://bitbucket.org/ericgazoni/openpyxl/wiki/Home
#  @see http://pypi.python.org/pypi/crontab/.11
#
#  @section LICENSE
#
#  Distributed under the terms of the MIT license.
#  @see http://de.wikipedia.org/wiki/MIT-Lizenz
#
__version__ = '$Id$'
#


import re, sys, os
import difflib
import BeautifulSoup
import urllib, StringIO, zipfile, csv
import mailbox, mimetypes, datetime, locale
import openpyxl.reader.excel
import crontab
import logging
import copy

import pagegenerators
import dtbext
# Splitting the bot into library parts
import wikipedia as pywikibot


bot_config = {    # unicode values
        'TemplateName':     u'User:DrTrigonBot/Subster',    # or 'template' for 'Flagged Revisions'
        'ErrorTemplate':    u'\n<noinclude>----\n<b>SubsterBot Exception (%s)</b>\n%s</noinclude>\n',

        # important to use a '.css' page here, since it HAS TO BE protected to
        # prevent malicious code injection !
        'ConfigCSSPage':    u'User:DrTrigon/DrTrigonBot/config.css',
        'CodeTemplate':     u'\n%s(DATA, *args)\n',
        'CRONMaxDelay':     1*24*60*60,       # bot runs daily

        # regex values
        'tag_regex':        re.compile('<.*?>', re.S | re.I),

        'var_regex_str':    u'<!--SUBSTER-%(var1)s-->%(cont)s<!--SUBSTER-%(var2)s-->',

        'mbox_file':         'mail_inbox',    # "drtrigon+subster@toolserver.org"
        'data_path':         '../data/subster',

        # bot paramater/options
        'param_default':    { 'url':   '',
                    'regex':           '',
                    'value':           '',
                    'count':           '0',
                    'notags':          '',
                    #'postproc':        '("","")',
                    'postproc':        '(\'\', \'\')',
                    'wiki':            'False',
                    'magicwords_only': 'False',
                    'beautifulsoup':   'False',        # DRTRIGON-88
                    'expandtemplates': 'False',        # DRTRIGON-93 (only with 'wiki')
                    'simple':          '',             # DRTRIGON-85
                    'zip':             'False',
                    'xlsx':            '',             #
                    'cron':            '',             # DRTRIGON-102
                    #'djvu': ... u"djvused -e 'n' \"%s\"" ... djvutext.py
                    #'pdf': ... u"pdftotext" or python module
                    #'imageocr', 'swfocr', ...
                    },

        'msg': {
            'de':  ( u'Bot: ',
                     u'substituiere %s tag(s).',
                   ),
            'en':  ( u'robot ',
                     u'substituting %s tag(s).',
                   ),
            'frr': ( u'Bot: ',
                     u'substituiere %s tag(s).',
                   ),
        },

        # this is a system parameter and should not be changed! (copy.deepcopy)
        'EditFlags':        {'minorEdit': True, 'botflag': True},

        # this is still VERY "HACKY" first approach to satisfy (Grip99)
        # http://de.wikipedia.org/wiki/Benutzer_Diskussion:DrTrigon#DrTrigonBot_ohne_Botflag
        # (may be a param 'boflag' should be added to template, that works only if enabled/listed here)
        'flagenable':       { u'Benutzer:Grip99/Wikipedia Diskussion:Vandalismusmeldung':
                                {'botflag': False}, # (enables to) disable botflag
                              u'Benutzer:Grip99/Wikipedia Diskussion:Sperrprüfung':
                                {'botflag': False}, # (enables to) disable botflag
                              u'Benutzer:Grip99/Wikipedia:Sperrprüfung':
                                {'botflag': False}, # (enables to) disable botflag
                            },


        # --- subster_irc.py; subster.bot_config ---
        # ...and second approach to satisfy (Grip99)
        # http://de.wikipedia.org/wiki/Benutzer_Diskussion:Grip99#Subster
        # may be use: Benutzer:DrTrigon/Benutzer:DrTrigonBot/config.css
        'difflink':         [ ( 'Wikipedia:Projektdiskussion/PRD-subst',  # target (with template)
                                'Wikipedia:Projektdiskussion',  # source (url in template)
                                {'subpages':  True,             # link params: ext. source with subpages?
                                 'flags':     {'minorEdit': False, 'botflag': False}, # link params: edit flags?
                                } ),
                            ],
}

## used/defined magic words, look also at bot_control
#  use, for example: '\<!--SUBSTER-BOTerror--\>\<!--SUBSTER-BOTerror--\>'
magic_words = {} # no magic word substitution (for empty dict)

# debug tools
# (look at 'bot_control.py' for more info)
debug = []                       # no write, all users
#debug.append( 'write2wiki' )    # write to wiki (operational mode)
#debug.append( 'code' )          # code debugging


class SubsterBot(dtbext.basic.BasicBot):
    '''
    Robot which will does substitutions of tags within wiki page content with external or
    other wiki text data. Like dynamic text updating.
    '''

    _param_default = bot_config['param_default']

    _tag_regex     = bot_config['tag_regex']
    _var_regex_str = bot_config['var_regex_str']%{'var1':'%(var)s','var2':'%(var)s','cont':'%(cont)s'}

    _BS_regex      = re.compile(u'(' + _var_regex_str%{'var':'BS:(.*?)','cont':'(.*?)'} + u')')
    # !!! access to full data (all attachements) should be possible too !!!
    #_BS_regex      = bot_config['var_regex_str']%{'var1':'%(var)s-BS:(.*?)','var2':'BS:(.*?)','cont':'(.*?)'}
    _BS_regex_str  = bot_config['var_regex_str']%{'var1':'BS:%(var)s','var2':'BS:/','cont':'%(cont)s'}

    # -template and subst-tag handling taken from MerlBot
    # -this bot could also be runned on my local wiki with an anacron-job

    def __init__(self):
        '''Constructor of SubsterBot(), initialize needed vars.'''

        pywikibot.output(u'\03{lightgreen}* Initialization of bot:\03{default}')

        dtbext.basic.BasicBot.__init__(self, bot_config, debug)

        # init constants
        self._userListPage = pywikibot.Page(self.site, bot_config['TemplateName'])
        self._ConfigPage   = pywikibot.Page(self.site, bot_config['ConfigCSSPage'])
        self.pagegen = pagegenerators.ReferringPageGenerator(self._userListPage, onlyTemplateInclusion=True)
        self._code   = self._ConfigPage.get()

    def run(self, sim=False, msg=bot_config['msg'], EditFlags=bot_config['EditFlags']):
        '''Run SubsterBot().'''

        pywikibot.output(u'\03{lightgreen}* Processing Template Backlink List:\03{default}')

        if sim:    self.pagegen = ['dummy']

        for page in self.pagegen:
            # setup source to get data from
            if sim:
                content = sim['content']
                params = [ sim ]
            else:
                pywikibot.output(u'Getting page "%s" via API from %s...'
                                 % (page.title(asLink=True), self.site))

                # get page content and operating mode
                content = self.load(page)
                params = self.loadTemplates(page, default=self._param_default)

            if not params: continue

            (substed_content, substed_tags) = self.subContent(content, params)

            # output result to page or return directly
            if sim:
                return substed_content
            else:
                # if changed, write!
                if (substed_content != content):
                #if substed_tags:
                    self.outputContentDiff(content, substed_content)

                    if 'write2wiki' not in debug:
                        pywikibot.output(u'\03{lightyellow}=== ! DEBUG MODE NOTHING WRITTEN TO WIKI ! ===\03{default}')
                        continue

                    head, mod = pywikibot.translate(self.site.lang, msg)
                    flags = copy.deepcopy(EditFlags)
                    if page.title() in bot_config['flagenable']:
                        flags.update( bot_config['flagenable'][page.title()] )
                    pywikibot.output(u'Flags used for writing: %s' % flags)
                    self.save(page, substed_content, head + mod % (", ".join(substed_tags)), **flags)
                else:
                    pywikibot.output(u'NOTHING TO DO!')

    def subContent(self, content, params):
        """Substitute the tags in content according to params.

           @param content: Content with tags to substitute.
           @type  content: string
           @param params: Params with data how to substitute tags.
           @type  params: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_content = content
        substed_tags = []  # DRTRIGON-73

        # 0.) subst (internal) magic words
        try:
            (substed_content, tags) = self.subBotMagicWords(substed_content)
            substed_tags += tags
        except:
            exc_info = sys.exc_info()
            (exception_only, result) = dtbext.pywikibot.gettraceback(exc_info)
            substed_content += bot_config['ErrorTemplate'] % ( dtbext.date.getTimeStmpNow(full=True, humanreadable=True, local=True), 
                                                               u' ' + result.replace(u'\n', u'\n ').rstrip() )
            substed_tags.append( u'>error:BotMagicWords<' )

        if (len(params) == 1) and eval(params[0]['magicwords_only']):
            return (substed_content, substed_tags)

        for item in params:
            # 1.) - 5.) subst templates
            try:
                (substed_content, tags) = self.subTemplate(substed_content, item)
                substed_tags += tags
            except:
                exc_info = sys.exc_info()
                (exception_only, result) = dtbext.pywikibot.gettraceback(exc_info)
                substed_content += bot_config['ErrorTemplate'] % ( dtbext.date.getTimeStmpNow(full=True, humanreadable=True, local=True), 
                                                                   u' ' + result.replace(u'\n', u'\n ').rstrip() )
                substed_tags.append( u'>error:Template<' )

        return (substed_content, substed_tags)

    def subBotMagicWords(self, content):
        """Substitute the DrTrigonBot Magic Word (tag)s in content.

           @param content: Content with tags to substitute.
           @type  content: string

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_tags = []  # DRTRIGON-73

        # 0.) subst (internal) magic words
        for subitem in magic_words.keys():
            prev_content = content
            content = self.get_var_regex(subitem).sub( (self._var_regex_str%{'var':subitem,'cont':magic_words[subitem]}),
                                                       content, 1)  # subst. once
            if (content != prev_content):
                substed_tags.append(subitem)

        return (content, substed_tags)

    def subTemplate(self, content, param):
        """Substitute the template tags in content according to param.

           @param content: Content with tags to substitute.
           @type  content: string
           @param param: Param with data how to substitute tags.
           @type  param: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_tags = []  # DRTRIGON-73
        prev_content = content

        # 0.2.) check cron/date
        if param['cron']:
            # [min] [hour] [day of month] [month] [day of week]
            # (date supported only, thus [min] and [hour] dropped)
            if not (param['cron'][0] == '@'):
                param['cron'] = '* * ' + param['cron']
            entry = crontab.CronTab(param['cron'])
            # find the delay from midnight (does not return 0.0 - but next)
            delay = entry.next(datetime.datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)-datetime.timedelta(microseconds=1))

            pywikibot.output(u'CRON delay for execution: %.3f (<= %i)' % (delay, bot_config['CRONMaxDelay']))

            if not (delay <= bot_config['CRONMaxDelay']):
                return (content, substed_tags)

        # 0.5.) check for 'simple' mode and get additional params
        if param['simple']:
            p = self.site.getExpandedString(param['simple'])
            param.update( pywikibot.extract_templates_and_params(p)[0][1] )

        # 1.) getUrl or wiki text
        # (security: check url not to point to a local file on the server,
        #  e.g. 'file://' - same as used in xsalt.py)
        secure = False
        for item in [u'http://', u'https://', u'mail://']:
            secure = secure or (param['url'][:len(item)] == item)
        param['wiki'] = eval(param['wiki'])
        param['zip']  = eval(param['zip'])
        if (not secure) and (not param['wiki']):
            return (content, substed_tags)
        if   param['wiki']:
            if eval(param['expandtemplates']): # DRTRIGON-93 (only with 'wiki')
                external_buffer = dtbext.pywikibot.Page(self.site, param['url']).get(expandtemplates=True)
            else:
                external_buffer = self.load( dtbext.pywikibot.Page(self.site, param['url']) )
        elif (param['url'][:7] == u'mail://'): # DRTRIGON-101
            mbox = SubsterMailbox(pywikibot.config.datafilepath(bot_config['data_path'], bot_config['mbox_file'], ''))
            external_buffer = mbox.find_data(param['url'])
            mbox.close()
        elif param['zip']:
            # !!! does zip deflate work with 'self.site.getUrl' ??!! (has to be made working!)
            external_buffer = urllib.urlopen(param['url']).read()
        else:
            external_buffer = self.site.getUrl(param['url'], no_hostname = True)

        # some intermediate processing (unzip, xlsx2csv, ...)
        if param['zip']:
            fileno          = 0 if (param['zip'] == True) else (param['zip']-1)
            external_buffer = self.unzip(external_buffer, fileno)
        if param['xlsx']:
            external_buffer = self.xlsx2csv(external_buffer, param['xlsx'])

        if not eval(param['beautifulsoup']):    # DRTRIGON-88
            # 2.) regexp
            #for subitem in param['regex']:
            subitem = param['regex']
            regex = re.compile(subitem, re.S | re.I)
            var_regex = self.get_var_regex(param['value'])

            # 3.) subst in content
            external_data = regex.search(external_buffer)

            if external_data:    # not None
                external_data = external_data.groups()

                pywikibot.output(u'Groups found by regex: %i' % len(external_data))

                if (len(external_data) == 1):
                    external_data = external_data[0]
                else:
                    external_data = str(external_data)
            logging.getLogger('subster').debug( external_data )

            if param['notags']:
                external_data = self._tag_regex.sub(param['notags'], external_data)
            logging.getLogger('subster').debug( external_data )

            # 4.) postprocessing
            param['postproc'] = eval(param['postproc'])
            func  = param['postproc'][0]    # needed by exec call of self._code
            DATA  = [ external_data ]       #
            args  = param['postproc'][1:]   #
            scope = {}                      # (scope to run in)
            scope.update( locals() )        # (add DATA, *args, ...)
            scope.update( globals() )       # (add imports and else)
            if func:
                exec(self._code + (bot_config['CodeTemplate'] % func), scope, scope)
                external_data = DATA[0]
            logging.getLogger('subster').debug( external_data )

            # 5.) subst content
            content = var_regex.sub((self._var_regex_str%{'var':param['value'],'cont':external_data}), content, int(param['count']))
            if (content != prev_content):
                substed_tags.append(param['value'])
        else:
            # DRTRIGON-88: Enable Beautiful Soup power for Subster
            BS_tags = self._BS_regex.findall(content)

            pywikibot.output(u'BeautifulSoup tags found by regex: %i' % len(BS_tags))

            for item in BS_tags:
                if not (item[3] == '/'): continue
                external_data = eval('BeautifulSoup.BeautifulSoup(external_buffer).%s' % item[1])
                content = content.replace(item[0], self._BS_regex_str%{'var':item[1],'cont':external_data}, 1)

            if (content != prev_content):
                substed_tags.append(u'BeautifulSoup')

        return (content, substed_tags)

    def outputContentDiff(self, content, substed_content):
        """Outputs the diff between the original and the new content.

           @param content: Original content.
           @type  content: string
           @param substed_content: New content.
           @type  substed_content: string

           Returns nothing, but outputs/prints the diff.
        """
        diff = difflib.Differ().compare(content.splitlines(1), substed_content.splitlines(1))
        diff = [ line for line in diff if line[0].strip() ]
        pywikibot.output(u'Diff:')
        pywikibot.output(u'--- ' * 15)
        pywikibot.output(u''.join(diff))
        pywikibot.output(u'--- ' * 15)

    def get_var_regex(self, var, cont='.*?'):
        """Get regex used/needed to find the tags to replace.

           @param var: The tag/variable name.
           @type  var: string
           @param cont: The content/value of the variable.
           @type  cont: string

           Return the according (and compiled) regex object.
        """
        return re.compile((self._var_regex_str%{'var':var,'cont':cont}), re.S | re.I)

    def unzip(self, external_buffer, i):
        """Convert zip data to plain format.
        """

        zip_buffer = zipfile.ZipFile(StringIO.StringIO(external_buffer))
        data_file  = zip_buffer.namelist()[i]
        external_buffer = zip_buffer.open(data_file).read().decode('latin-1')

        return external_buffer

    def xlsx2csv(self, external_buffer, sheet):
        """Convert xlsx (EXCEL) data to csv format.
        """

        wb = openpyxl.reader.excel.load_workbook(StringIO.StringIO(external_buffer), use_iterators = True)

        sheet_ranges = wb.get_sheet_by_name(name = sheet)

        output = StringIO.StringIO()
        spamWriter = csv.writer(output)

        for row in sheet_ranges.iter_rows(): # it brings a new method: iter_rows()
            spamWriter.writerow([ cell.internal_value for cell in row ])

        external_buffer = output.getvalue()
        output.close()

        return external_buffer


class SubsterMailbox(mailbox.mbox):
    def __init__(self, mbox_file):
        mailbox.mbox.__init__(self, mbox_file)
        self.lock()

        self.remove_duplicates()

    def remove_duplicates(self):
        """Find mails with same 'From' (sender) and remove all
        except the most recent one.
        """

        unique = {}
        remove = []
        for i, message in enumerate(self):
            sender   = message['from']       # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            locale.setlocale(locale.LC_TIME, 'en_US')   # datetime in mails saved in 'en_US' by toolserver
            timestmp = re.split('[+-]', timestmp)[0].strip()
            timestmp = datetime.datetime.strptime(timestmp, '%a, %d %b %Y %H:%M:%S')

            if sender in unique:
                (j, timestmp_j) = unique[sender]

                if (timestmp >= timestmp_j): 
                    remove.append( j )
                else:
                    remove.append( i )
            else:
                unique[sender] = (i, timestmp)

        remove.reverse()
        for i in remove:
            self.remove(i)

        self.flush()
        #self.close()

        if remove:
            pywikibot.output('Removed %i depreciated email data source(s).' % len(remove))

    def find_data(self, url):
        """Find mail according to given 'From' (sender).
        """

        url = (url[:7], ) + tuple(url[7:].split('/'))
        content = ''

        for i, message in enumerate(self):
            sender   = message['from']          # Could possibly be None.
            subject  = message['subject']       # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            if sender and url[1] in sender:
                # data found
                pywikibot.output('Found email data source:')
                pywikibot.output('%i / %s / %s / %s' % (i, sender, subject, timestmp))

                full = (url[2] == 'attachment-full')
                if   (url[2] == 'all'):
                    content = message.as_string(True)
                elif (url[2] == 'attachment') or full:
                    counter = 1
                    content = ''
                    for part in message.walk():
                        # multipart/* are just containers
                        if part.get_content_maintype() == 'multipart':
                            continue
                        # Applications should really sanitize the given filename so that an
                        # email message can't be used to overwrite important files
                        filename = part.get_filename()
                        if filename or full:
                            if not filename:
                                ext = mimetypes.guess_extension(part.get_content_type())
                                if not ext:
                                    # Use a generic bag-of-bits extension
                                    ext = '.bin'
                                filename = 'part-%03d%s' % (counter, ext)
                            counter += 1

                            content += part.get_payload(decode=True)
                            pywikibot.output('Found attachment: "' + filename + '"')

                            if not full: break

                    break

        return content


def main():
    args = pywikibot.handleArgs()
    bot  = SubsterBot()   # for several user's, but what about complete automation (continous running...)
    if len(args) > 0:
        for arg in args:
            if arg[:2] == "u'": arg = eval(arg)        # for 'bot_control.py' and unicode compatibility
            if   (arg[:5] == "-auto") \
                 or (arg[:5] == "-cron"):
                bot.silent = True
            elif (arg == "-all") \
                 or (arg == "-default") \
                 or ("-subster" in arg):
                pass
            elif (arg == "-no_magic_words"):
                pass
            else:
                pywikibot.showHelp()
                return
    try:
        bot.run()
    except KeyboardInterrupt:
        pywikibot.output('\nQuitting program...')

if __name__ == "__main__":
    try:
        main()
    finally:
        pywikibot.stopme()

